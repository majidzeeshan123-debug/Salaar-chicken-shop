"""
Database module for Poultry Shop Management App
Uses Excel (openpyxl) as a local file-based database - works well with Pydroid.
"""
import os
from openpyxl import Workbook, load_workbook

DB_FILE = "poultry_shop_data.xlsx"

SHEETS = {
    "Purchases": ["Date", "Supplier", "Weight_KG", "Rate_Per_KG", "Total_Cost", "Notes"],
    "Processing": ["Date", "Live_Weight_Used_KG", "Meat_Weight_KG",
                   "KalejiPota_Weight_KG", "Waste_Weight_KG", "Notes"],
    "Sales": ["Date", "Category", "Weight_KG", "Rate_Per_KG", "Total_Amount",
              "Customer", "Payment_Type", "Notes"],
    "Expenses": ["Date", "Category", "Amount", "Notes"],
    "Khata": ["Date", "Customer", "Phone", "Type", "Amount", "Balance", "Notes"],
}

CATEGORIES = ["Live Chicken", "Meat", "Kaleji-Pota", "Waste"]


class Database:
    def __init__(self, path=None):
        self.path = path or DB_FILE
        if not os.path.exists(self.path):
            self._create_new_db()
        self.wb = load_workbook(self.path)
        # Make sure any missing sheets get added (for future upgrades)
        changed = False
        for sheet_name, headers in SHEETS.items():
            if sheet_name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(sheet_name)
                ws.append(headers)
                changed = True
        if changed:
            self._save()

    def _create_new_db(self):
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, headers in SHEETS.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        wb.save(self.path)

    def _save(self):
        self.wb.save(self.path)

    def add_row(self, sheet_name, row):
        ws = self.wb[sheet_name]
        ws.append(row)
        self._save()

    def get_all_rows(self, sheet_name):
        ws = self.wb[sheet_name]
        return list(ws.iter_rows(min_row=2, values_only=True))

    # ---------- Purchases (buying live chicken) ----------
    def add_purchase(self, date, supplier, weight_kg, rate_per_kg, notes=""):
        weight_kg = float(weight_kg)
        rate_per_kg = float(rate_per_kg)
        total_cost = round(weight_kg * rate_per_kg, 2)
        self.add_row("Purchases", [date, supplier, weight_kg, rate_per_kg, total_cost, notes])
        return total_cost

    # ---------- Processing (cutting live chicken into meat / kaleji-pota / waste) ----------
    def add_processing(self, date, live_weight_used, meat_weight, kaleji_pota_weight, waste_weight, notes=""):
        live_weight_used = float(live_weight_used)
        meat_weight = float(meat_weight)
        kaleji_pota_weight = float(kaleji_pota_weight)
        waste_weight = float(waste_weight)
        self.add_row("Processing", [date, live_weight_used, meat_weight,
                                     kaleji_pota_weight, waste_weight, notes])

    # ---------- Sales ----------
    def add_sale(self, date, category, weight_kg, rate_per_kg, customer, payment_type, notes=""):
        weight_kg = float(weight_kg)
        rate_per_kg = float(rate_per_kg)
        total_amount = round(weight_kg * rate_per_kg, 2)
        self.add_row("Sales", [date, category, weight_kg, rate_per_kg,
                                total_amount, customer, payment_type, notes])
        if payment_type == "Credit":
            self.add_khata_entry(date, customer, "", "Credit Sale", total_amount,
                                  notes=f"{category} sale on credit")
        return total_amount

    # ---------- Expenses ----------
    def add_expense(self, date, category, amount, notes=""):
        amount = float(amount)
        self.add_row("Expenses", [date, category, amount, notes])

    # ---------- Khata (credit ledger) ----------
    def add_khata_entry(self, date, customer, phone, entry_type, amount, notes=""):
        amount = float(amount)
        balance = self.get_khata_balance(customer)
        if entry_type == "Credit Sale":
            balance += amount
        elif entry_type == "Payment Received":
            balance -= amount
        self.add_row("Khata", [date, customer, phone, entry_type, amount, round(balance, 2), notes])
        return balance

    def get_khata_balance(self, customer):
        balance = 0
        for r in self.get_all_rows("Khata"):
            if r[1] == customer:
                balance = r[5]
        return balance or 0

    def get_khata_customers(self):
        customers = {}
        for r in self.get_all_rows("Khata"):
            customers[r[1]] = r[5]
        return {k: v for k, v in customers.items() if v}

    # ---------- Stock ----------
    def get_stock_summary(self):
        purchases = self.get_all_rows("Purchases")
        processing = self.get_all_rows("Processing")
        sales = self.get_all_rows("Sales")

        total_purchased = sum(r[2] for r in purchases if r[2])
        live_used = sum(r[1] for r in processing if r[1])
        meat_produced = sum(r[2] for r in processing if r[2])
        kalejipota_produced = sum(r[3] for r in processing if r[3])
        waste_produced = sum(r[4] for r in processing if r[4])

        def sold(category):
            return sum(r[2] for r in sales if r[1] == category and r[2])

        return {
            "Live Chicken (KG)": round(total_purchased - live_used - sold("Live Chicken"), 2),
            "Meat (KG)": round(meat_produced - sold("Meat"), 2),
            "Kaleji-Pota (KG)": round(kalejipota_produced - sold("Kaleji-Pota"), 2),
            "Waste (KG)": round(waste_produced - sold("Waste"), 2),
        }

    # ---------- Daily / date-range summary ----------
    def get_summary_for_date(self, date_str):
        purchases = [r for r in self.get_all_rows("Purchases") if str(r[0]) == date_str]
        sales = [r for r in self.get_all_rows("Sales") if str(r[0]) == date_str]
        expenses = [r for r in self.get_all_rows("Expenses") if str(r[0]) == date_str]

        total_purchase = sum(r[4] for r in purchases if r[4])
        total_sales = sum(r[4] for r in sales if r[4])
        total_expense = sum(r[2] for r in expenses if r[2])
        net_profit = round(total_sales - total_purchase - total_expense, 2)

        sales_by_category = {}
        for r in sales:
            sales_by_category[r[1]] = sales_by_category.get(r[1], 0) + (r[4] or 0)

        return {
            "Total Purchase": round(total_purchase, 2),
            "Total Sales": round(total_sales, 2),
            "Total Expense": round(total_expense, 2),
            "Net Profit": net_profit,
            "Sales By Category": sales_by_category,
        }
